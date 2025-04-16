import pandas as pd
import numpy as np
import math
import os
from collections import Counter
from openpyxl import load_workbook
from joblib import Parallel, delayed
import gc

def infer_column_types(df):
    col_types = []
    for col in df.columns:
        values = df[col].dropna()
        if values.empty:
            col_types.append('categorical')
            continue
        
        coerced = pd.to_numeric(values, errors='coerce')
        if coerced.isna().sum() == 0:
            col_types.append('numerical')
        else:
            col_types.append('categorical')
    return col_types

def gray_knn_impute_mixed_vectorized(df, k=5, xi=0.5, batch_size=50, parallel=False, n_jobs=4):
    df_filled = df.copy()
    df_filled.replace(r'^\s*$', np.nan, regex=True, inplace=True)
    data = df_filled.to_numpy(dtype=object)
    col_types = infer_column_types(df_filled)
    print(f"Column types for {df.shape}: {col_types}", flush=True)

    def impute_column(j, data, k, xi, col_types, batch_size):
        col = data[:, j]
        missing = np.array([pd.isna(v) for v in col])
        if not np.any(missing):
            return None

        missing_rows_idx = np.where(missing)[0]
        missing_rows = data[missing_rows_idx, :]
        neighbor_mask = ~missing
        neighbors = data[neighbor_mask, :]
        if len(neighbors) < k:
            if col_types[j] == 'numerical':
                column_mean = np.nanmean(pd.to_numeric(col, errors='coerce'))
                value = column_mean if not np.isnan(column_mean) else 0.0
            else:
                mode_value = pd.Series(col).mode()
                value = mode_value[0] if not mode_value.empty else np.nan
            return [(idx, j, value) for idx in missing_rows_idx]

        n_missing = missing_rows.shape[0]
        batch_size = min(batch_size, max(10, int(n_missing / 10)))
        grg_all = np.zeros((n_missing, neighbors.shape[0]), dtype=np.float32)

        for start in range(0, n_missing, batch_size):
            end = min(start + batch_size, n_missing)
            batch_missing_rows = missing_rows[start:end, :]
            batch_size_actual = end - start

            grc_sum = np.zeros((batch_size_actual, neighbors.shape[0]), dtype=np.float32)
            valid_cols = 0

            for col_idx, col_type in enumerate(col_types):
                delta = np.zeros((batch_size_actual, neighbors.shape[0]), dtype=np.float32)
                if col_type == 'numerical':
                    try:
                        m_col = pd.to_numeric(neighbors[:, col_idx], errors='coerce')
                        q_col = pd.to_numeric(batch_missing_rows[:, col_idx], errors='coerce')
                        if np.any(pd.isna(m_col)) or np.any(pd.isna(q_col)):
                            delta[:, :] = (neighbors[:, col_idx] != batch_missing_rows[:, col_idx, np.newaxis]).astype(np.float32)
                        else:
                            delta[:, :] = np.abs(m_col - q_col[:, np.newaxis])
                    except Exception as e:
                        print(f"Error in column {col_idx} (numerical): {str(e)}", flush=True)
                        delta[:, :] = (neighbors[:, col_idx] != batch_missing_rows[:, col_idx, np.newaxis]).astype(np.float32)
                else:
                    try:
                        delta[:, :] = (neighbors[:, col_idx] != batch_missing_rows[:, col_idx, np.newaxis]).astype(np.float32)
                        m_isna = np.array([pd.isna(v) for v in neighbors[:, col_idx]])
                        q_isna = np.array([pd.isna(v) for v in batch_missing_rows[:, col_idx]])
                        delta[m_isna | q_isna[:, np.newaxis]] = np.nan
                    except Exception as e:
                        print(f"Error in column {col_idx} (categorical): {str(e)}", flush=True)
                        delta[:, :] = np.nan

                min_delta, max_delta = np.nanmin(delta), np.nanmax(delta)
                if np.isnan(max_delta) or max_delta == min_delta:
                    grc = np.ones_like(delta, dtype=np.float32) if min_delta == 0 else np.zeros_like(delta, dtype=np.float32)
                else:
                    grc = (min_delta + xi * max_delta) / (delta + xi * max_delta)
                    grc = np.nan_to_num(grc, nan=0.0)
                grc_sum += grc
                if not np.all(np.isnan(delta)):
                    valid_cols += 1

            if valid_cols > 0:
                grg_all[start:end, :] = grc_sum / valid_cols
            else:
                grg_all[start:end, :] = 0.0

        top_k_indices = np.argsort(grg_all, axis=1)[:, -k:]
        result = []
        for i, row_idx in enumerate(missing_rows_idx):
            top_k_vals = neighbors[top_k_indices[i], j]
            top_k_weights = grg_all[i, top_k_indices[i]]
            weight_sum = np.nansum(top_k_weights)
            if weight_sum > 0:
                if col_types[j] == 'numerical':
                    valid_pairs = [(float(v), w) for v, w in zip(top_k_vals, top_k_weights) 
                                 if pd.notna(v) and str(v).replace('.', '').replace('-', '').replace('e', '').isdigit()]
                    if valid_pairs:
                        valid_vals, valid_weights = zip(*valid_pairs)
                        if sum(valid_weights) > 0:
                            value = np.sum([v * w for v, w in zip(valid_vals, valid_weights)]) / sum(valid_weights)
                        else:
                            column_mean = np.nanmean(pd.to_numeric(col, errors='coerce'))
                            value = column_mean if not np.isnan(column_mean) else 0.0
                    else:
                        column_mean = np.nanmean(pd.to_numeric(col, errors='coerce'))
                        value = column_mean if not np.isnan(column_mean) else 0.0
                else:
                    weighted_values = Counter()
                    for val, weight in zip(top_k_vals, top_k_weights):
                        if pd.notna(val):
                            weighted_values[val] += weight
                    value = weighted_values.most_common(1)[0][0] if weighted_values else pd.Series(col).mode()[0] if not pd.Series(col).mode().empty else np.nan
            else:
                if col_types[j] == 'numerical':
                    column_mean = np.nanmean(pd.to_numeric(col, errors='coerce'))
                    value = column_mean if not np.isnan(column_mean) else 0.0
                else:
                    mode_value = pd.Series(col).mode()
                    value = mode_value[0] if not mode_value.empty else np.nan
            result.append((row_idx, j, value))
        return result

    if parallel:
        results = Parallel(n_jobs=n_jobs)(delayed(impute_column)(j, data, k, xi, col_types, batch_size) for j in range(data.shape[1]))
        for result in results:
            if result is None:
                continue
            for row_idx, col_idx, value in result:
                data[row_idx, col_idx] = value
    else:
        for j in range(data.shape[1]):
            result = impute_column(j, data, k, xi, col_types, batch_size)
            if result is None:
                continue
            for row_idx, col_idx, value in result:
                data[row_idx, col_idx] = value

    df_filled.iloc[:, :] = data
    return df_filled

def gray_knn_impute_gpu(df, k=5, xi=0.5, batch_size=50, parallel=False, n_jobs=4):
    df_filled = df.copy()
    df_filled.replace(r'^\s*$', np.nan, regex=True, inplace=True)
    data = df_filled.to_numpy(dtype=np.float32)

    def impute_column(j, data, k, xi, batch_size):
        col = data[:, j]
        missing = np.isnan(col)
        if not np.any(missing):
            return None

        missing_rows_idx = np.where(missing)[0]
        missing_rows = data[missing_rows_idx, :]
        neighbor_mask = ~missing
        neighbors = data[neighbor_mask, :]
        if len(neighbors) < k:
            column_mean = np.nanmean(col)
            value = column_mean if not np.isnan(column_mean) else 0.0
            return [(idx, j, value) for idx in missing_rows_idx]

        n_missing = missing_rows.shape[0]
        batch_size = min(batch_size, max(10, int(n_missing / 10)))
        grg_all = np.zeros((n_missing, neighbors.shape[0]), dtype=np.float32)

        for start in range(0, n_missing, batch_size):
            end = min(start + batch_size, n_missing)
            batch_missing_rows = missing_rows[start:end, :]
            batch_size_actual = end - start

            grc_sum = np.zeros((batch_size_actual, neighbors.shape[0]), dtype=np.float32)
            valid_cols = 0

            for col_idx in range(data.shape[1]):
                delta = np.zeros((batch_size_actual, neighbors.shape[0]), dtype=np.float32)
                m_col = neighbors[:, col_idx]
                q_col = batch_missing_rows[:, col_idx]
                delta[:, :] = np.abs(m_col - q_col[:, np.newaxis])
                delta[np.isnan(delta)] = np.nan

                min_delta, max_delta = np.nanmin(delta), np.nanmax(delta)
                if np.isnan(max_delta) or max_delta == min_delta:
                    grc = np.ones_like(delta, dtype=np.float32) if min_delta == 0 else np.zeros_like(delta, dtype=np.float32)
                else:
                    grc = (min_delta + xi * max_delta) / (delta + xi * max_delta)
                    grc = np.nan_to_num(grc, nan=0.0)
                grc_sum += grc
                if not np.all(np.isnan(delta)):
                    valid_cols += 1

            if valid_cols > 0:
                grg_all[start:end, :] = grc_sum / valid_cols
            else:
                grg_all[start:end, :] = 0.0

        top_k_indices = np.argsort(grg_all, axis=1)[:, -k:]
        result = []
        for i, row_idx in enumerate(missing_rows_idx):
            top_k_vals = neighbors[top_k_indices[i], j]
            top_k_weights = grg_all[i, top_k_indices[i]]
            weight_sum = np.nansum(top_k_weights)
            if weight_sum > 0:
                valid_pairs = [(float(v), w) for v, w in zip(top_k_vals, top_k_weights) if not np.isnan(v)]
                if valid_pairs:
                    valid_vals, valid_weights = zip(*valid_pairs)
                    if sum(valid_weights) > 0:
                        value = np.sum([v * w for v, w in zip(valid_vals, valid_weights)]) / sum(valid_weights)
                    else:
                        column_mean = np.nanmean(col)
                        value = column_mean if not np.isnan(column_mean) else 0.0
                else:
                    column_mean = np.nanmean(col)
                    value = column_mean if not np.isnan(column_mean) else 0.0
            else:
                column_mean = np.nanmean(col)
                value = column_mean if not np.isnan(column_mean) else 0.0
            result.append((row_idx, j, value))
        return result

    if parallel:
        results = Parallel(n_jobs=n_jobs)(delayed(impute_column)(j, data, k, xi, batch_size) for j in range(data.shape[1]))
        for result in results:
            if result is None:
                continue
            for row_idx, col_idx, value in result:
                data[row_idx, col_idx] = value
    else:
        for j in range(data.shape[1]):
            result = impute_column(j, data, k, xi, batch_size)
            if result is None:
                continue
            for row_idx, col_idx, value in result:
                data[row_idx, col_idx] = value

    df_filled.iloc[:, :] = data
    return df_filled

def gray_knn_impute_cat_vectorized(df, k=5, xi=0.9, batch_size=50, parallel=False, n_jobs=4):
    df_filled = df.copy()
    df_filled.replace(r'^\s*$', np.nan, regex=True, inplace=True)
    data = df_filled.to_numpy(dtype=object)

    def impute_column(j, data, k, xi, batch_size):
        col = data[:, j]
        missing = np.array([pd.isna(v) for v in col])
        if not np.any(missing):
            return None

        missing_rows_idx = np.where(missing)[0]
        missing_rows = data[missing_rows_idx, :]
        neighbor_mask = ~missing
        neighbors = data[neighbor_mask, :]
        if len(neighbors) < k:
            mode_value = pd.Series(col).mode()
            value = mode_value[0] if not mode_value.empty else np.nan
            return [(idx, j, value) for idx in missing_rows_idx]

        n_missing = missing_rows.shape[0]
        batch_size = min(batch_size, max(10, int(n_missing / 10)))
        grg_all = np.zeros((n_missing, neighbors.shape[0]), dtype=np.float32)

        for start in range(0, n_missing, batch_size):
            end = min(start + batch_size, n_missing)
            batch_missing_rows = missing_rows[start:end, :]
            batch_size_actual = end - start

            grc_sum = np.zeros((batch_size_actual, neighbors.shape[0]), dtype=np.float32)
            valid_cols = 0

            for col_idx in range(data.shape[1]):
                delta = np.zeros((batch_size_actual, neighbors.shape[0]), dtype=np.float32)
                delta[:, :] = (neighbors[:, col_idx] != batch_missing_rows[:, col_idx, np.newaxis]).astype(np.float32)
                m_isna = np.array([pd.isna(v) for v in neighbors[:, col_idx]])
                q_isna = np.array([pd.isna(v) for v in batch_missing_rows[:, col_idx]])
                delta[m_isna | q_isna[:, np.newaxis]] = np.nan

                min_delta, max_delta = np.nanmin(delta), np.nanmax(delta)
                if np.isnan(max_delta) or max_delta == min_delta:
                    grc = np.ones_like(delta, dtype=np.float32) if min_delta == 0 else np.zeros_like(delta, dtype=np.float32)
                else:
                    grc = (min_delta + xi * max_delta) / (delta + xi * max_delta)
                    grc = np.nan_to_num(grc, nan=0.0)
                grc_sum += grc
                if not np.all(np.isnan(delta)):
                    valid_cols += 1

            if valid_cols > 0:
                grg_all[start:end, :] = grc_sum / valid_cols
            else:
                grg_all[start:end, :] = 0.0

        top_k_indices = np.argsort(grg_all, axis=1)[:, -k:]
        result = []
        for i, row_idx in enumerate(missing_rows_idx):
            top_k_vals = neighbors[top_k_indices[i], j]
            top_k_weights = grg_all[i, top_k_indices[i]]
            weight_sum = np.nansum(top_k_weights)
            if weight_sum > 0:
                weighted_values = Counter()
                for val, weight in zip(top_k_vals, top_k_weights):
                    if pd.notna(val):
                        weighted_values[val] += weight
                value = weighted_values.most_common(1)[0][0] if weighted_values else pd.Series(col).mode()[0] if not pd.Series(col).mode().empty else np.nan
            else:
                mode_value = pd.Series(col).mode()
                value = mode_value[0] if not mode_value.empty else np.nan
            result.append((row_idx, j, value))
        return result

    if parallel:
        results = Parallel(n_jobs=n_jobs)(delayed(impute_column)(j, data, k, xi, batch_size) for j in range(data.shape[1]))
        for result in results:
            if result is None:
                continue
            for row_idx, col_idx, value in result:
                data[row_idx, col_idx] = value
    else:
        for j in range(data.shape[1]):
            result = impute_column(j, data, k, xi, batch_size)
            if result is None:
                continue
            for row_idx, col_idx, value in result:
                data[row_idx, col_idx] = value

    df_filled.iloc[:, :] = data
    return df_filled

def process_large_datasets(incomplete_root, original_root, imputed_root, output_excel="Table-NRMS-AE.xlsx"):
    large_subfolders = [
        ('Letter', 'numerical', 19999),
        ('Adult', 'mixed', 45221),
        ('C4', 'categorical', 67556),
        ('KDD', 'mixed', 494019)
    ]
    
    dataset_types = {
        'numerical': (gray_knn_impute_gpu, {'xi': 0.5}),
        'categorical': (gray_knn_impute_cat_vectorized, {'xi': 0.9}),
        'mixed': (gray_knn_impute_mixed_vectorized, {'xi': 0.5})
    }
    
    if not os.path.exists(imputed_root):
        os.makedirs(imputed_root)
    
    try:
        wb = load_workbook(output_excel)
        ws = wb.active
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file {output_excel} not found.")
    
    print("Processing large datasets: Letter, Adult, C4, KDD", flush=True)
    
    for subfolder, data_type, rows in large_subfolders:
        k = max(50, int(math.sqrt(rows) / 2)) if data_type == 'categorical' else max(5, int(math.sqrt(rows) / 2))
        
        print(f"Subfolder: {subfolder}, Rows: {rows}, Using k: {k}", flush=True)
        
        subfolder_path = os.path.join(incomplete_root, subfolder)
        if not os.path.isdir(subfolder_path):
            print(f"Subfolder not found: {subfolder_path}", flush=True)
            continue
        
        full_file = os.path.join(original_root, f"{subfolder}.xlsx")
        imputed_subfolder = os.path.join(imputed_root, subfolder)
        
        if not os.path.exists(imputed_subfolder):
            os.makedirs(imputed_subfolder)
        
        print(f"Checking file: {full_file}", flush=True)
        if not os.path.exists(full_file):
            print(f"Full file not found: {full_file}", flush=True)
            continue
        
        dtype = None if data_type == 'numerical' else str
        try:
            df_true = pd.read_excel(full_file, header=None, dtype=dtype)
            print(f"Successfully loaded df_true for {subfolder}", flush=True)
        except Exception as e:
            print(f"Failed to load df_true for {subfolder}: {str(e)}", flush=True)
            continue

        if df_true is None or df_true.empty:
            print(f"Skipping {subfolder} as df_true is empty or missing.", flush=True)
            continue
        
        impute_func, params = dataset_types[data_type]
        params['k'] = k
        params['parallel'] = True
        
        for file in os.listdir(subfolder_path):
            if not file.endswith('.xlsx'):
                continue
            
            print(f"Processing file: {file}", flush=True)
            missing_file = os.path.join(subfolder_path, file)
            df_missing = pd.read_excel(missing_file, header=None, dtype=dtype)
            missing_mask = df_missing.isna()
            
            if 'batch_size' in params:
                df_imputed = impute_func(df_missing, **params)
            else:
                df_imputed = impute_func(df_missing, **params)
            
            if df_imputed.shape != df_true.shape:
                print(f"Shape mismatch: {file} - df_imputed: {df_imputed.shape}, df_true: {df_true.shape}", flush=True)
                continue
            
            if 'df_true' not in locals():
                print(f"df_true not defined for subfolder {subfolder}, file {file}", flush=True)
                continue

            nrms = 'N/A'
            ae = 'N/A'
            total_missing_cat = 0
            key_params = f"k={k}, xi={params.get('xi', 'N/A')}"
            
            col_types = infer_column_types(df_missing)
            numerical_cols = [i for i, t in enumerate(col_types) if t == 'numerical']
            categorical_cols = [i for i, t in enumerate(col_types) if t == 'categorical']

            if numerical_cols:
                mask = missing_mask.to_numpy()
                num_mask = mask[:, numerical_cols]
                try:
                    imputed_numeric = df_imputed.iloc[:, numerical_cols].apply(lambda x: pd.to_numeric(x, errors='coerce')).to_numpy()
                    true_numeric = df_true.iloc[:, numerical_cols].apply(lambda x: pd.to_numeric(x, errors='coerce')).to_numpy()
                    diff = imputed_numeric[num_mask] - true_numeric[num_mask]
                    nrms_num = np.sum(diff ** 2)
                    nrms_denom = np.sum(true_numeric[num_mask] ** 2)
                    nrms = np.sqrt(nrms_num) / np.sqrt(nrms_denom) if nrms_denom > 0 else 0
                    print(f"NRMS for {file}: {nrms}", flush=True)
                except Exception as e:
                    print(f"Error computing NRMS for {file}: {str(e)}", flush=True)
            
            if categorical_cols:
                mask = missing_mask.to_numpy()
                cat_mask = mask[:, categorical_cols]
                try:
                    imputed_values = df_imputed.iloc[:, categorical_cols].to_numpy()[cat_mask]
                    true_values = df_true.iloc[:, categorical_cols].to_numpy()[cat_mask]
                    correct_matches = np.sum((imputed_values == true_values) & cat_mask[cat_mask])
                    total_missing_cat = np.sum(cat_mask)
                    ae = correct_matches / total_missing_cat if total_missing_cat > 0 else 0
                    print(f"AE for {file}: {ae}", flush=True)
                except Exception as e:
                    print(f"Error computing AE for {file}: {str(e)}", flush=True)

            dataset_name = file.split('.')[0]
            
            updated = False
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
                if row[0].value == dataset_name:
                    try:
                        row[1].value = nrms if nrms != 'N/A' else ''
                        row[2].value = ae if ae != 'N/A' else ''
                        row[3].value = key_params
                        updated = True
                        wb.save(output_excel)
                        print(f"Updated Excel row for {dataset_name}", flush=True)
                    except PermissionError:
                        print(f"Error: Permission denied when updating row {row[0].row} for {dataset_name}. Close the excel file", flush=True)
                    except Exception as e:
                        print(f"Error updating row {row[0].row} for {dataset_name}: {str(e)}", flush=True)
                    break

            if not updated:
                print(f"Warning: Dataset '{dataset_name}' not found in Excel file.", flush=True)
            
            imputed_file = os.path.join(imputed_subfolder, f"imputed_{file}")
            df_imputed.to_excel(imputed_file, index=False, header=False)
            print(f"Saved imputed file: {imputed_file}", flush=True)
            
            del df_missing, df_imputed
            gc.collect()
    
    try:
        wb.save(output_excel)
        print(f"Excel file successfully updated: {output_excel}", flush=True)
    except PermissionError:
        raise Exception(f"Error: Permission denied when saving {output_excel}.")
    except Exception as e:
        raise Exception(f"Error saving Excel file: {str(e)}")

# Run for all datasets
incomplete_root = r"Incomplete Datasets Without Labels"
original_root = r"Original Datasets Without Labels"
imputed_root = r"Imputed Datasets"
process_large_datasets(incomplete_root, original_root, imputed_root, output_excel="Table-NRMS-AE.xlsx")