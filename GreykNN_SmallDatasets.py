import pandas as pd
import numpy as np
import cupy as cp
from collections import Counter
import os
import math
from openpyxl import load_workbook

def gray_knn_impute_gpu(df, k=5, xi=0.5):
    df_filled = df.copy()
    df_filled = df_filled.apply(pd.to_numeric, errors='coerce')
    numeric_data = df_filled.to_numpy()
    if not np.issubdtype(numeric_data.dtype, np.number):
        raise ValueError(f"Non-numeric data detected: {numeric_data.dtype}")
    data_gpu = cp.asarray(numeric_data)
    
    for j in range(data_gpu.shape[1]):
        missing_mask = cp.isnan(data_gpu[:, j])
        if not cp.any(missing_mask):
            print(f"Column {j}: No missing values, skipping.", flush=True)
            continue
        missing_rows_idx = cp.where(missing_mask)[0]
        missing_rows = data_gpu[missing_rows_idx, :]
        neighbor_mask = ~cp.isnan(data_gpu[:, j])
        neighbors = data_gpu[neighbor_mask, :]
        print(f"Column {j}: {len(neighbors)} non-missing neighbors found (k={k}).", flush=True)
        k_adjusted = min(k, len(neighbors))
        if k_adjusted == 0:
            print(f"Column {j}: No neighbors available, using column mean.", flush=True)
            column_mean = np.nanmean(numeric_data[:, j])
            numeric_data[missing_rows_idx.get(), j] = column_mean
            continue
        grg = gray_relational_grade_gpu(neighbors, missing_rows, xi)
        top_k_indices = cp.argsort(grg, axis=1)[:, -k_adjusted:]
        for i, row_idx in enumerate(missing_rows_idx.get()):
            top_k_vals = neighbors[top_k_indices[i], j].get()
            top_k_weights = grg[i, top_k_indices[i]].get()
            weight_sum = cp.sum(top_k_weights)
            print(f"Row {row_idx}, Column {j}: Sum of top-k weights = {weight_sum}", flush=True)
            if weight_sum > 0:
                numeric_data[row_idx, j] = cp.sum(top_k_vals * top_k_weights) / weight_sum
            else:
                print(f"Row {row_idx}, Column {j}: Weight sum <= 0 ({weight_sum}), using column mean.", flush=True)
                column_mean = np.nanmean(numeric_data[:, j])
                numeric_data[row_idx, j] = column_mean if not np.isnan(column_mean) else 0.0
    
    df_filled.iloc[:, :] = numeric_data
    return df_filled

def gray_relational_grade_gpu(matrix, queries, xi=0.5):
    matrix_gpu = cp.asarray(matrix)
    queries_gpu = cp.asarray(queries)
    delta = cp.abs(matrix_gpu - queries_gpu[:, cp.newaxis, :])
    min_delta, max_delta = cp.nanmin(delta), cp.nanmax(delta)
    
    if cp.isnan(max_delta) or max_delta == min_delta:
        grc = cp.ones_like(delta, dtype=cp.float32) if min_delta == 0 else cp.zeros_like(delta, dtype=cp.float32)
    else:
        grc = (min_delta + xi * max_delta) / (delta + xi * max_delta)
        grc = cp.nan_to_num(grc, nan=0.0)
    
    return cp.nanmean(grc, axis=2)

def gray_knn_impute_cat_vectorized(df, k=50, xi=0.9):
    df_filled = df.copy()
    df_filled.replace(r'^\s*$', np.nan, regex=True, inplace=True)
    data = df_filled.to_numpy(dtype=object)
    
    for j in range(data.shape[1]):
        missing_mask = pd.isna(data[:, j])
        if not missing_mask.any():
            print(f"Column {j}: No missing values, skipping.", flush=True)
            continue
        missing_rows_idx = np.where(missing_mask)[0]
        missing_rows = data[missing_rows_idx, :]
        neighbor_mask = ~pd.isna(data[:, j])
        neighbors = data[neighbor_mask, :]
        if len(neighbors) < k:
            print(f"Column {j}: Insufficient neighbors ({len(neighbors)} < {k}), using mode.", flush=True)
            mode_value = df_filled[j].mode()[0] if not df_filled[j].mode().empty else np.nan
            data[missing_rows_idx, j] = mode_value
            continue
        k_adjusted = min(k, len(neighbors))
        grg = gray_relational_grade_cat_vectorized(neighbors, missing_rows, xi)
        top_k_indices = np.argsort(grg, axis=1)[:, -k_adjusted:]
        for i, row_idx in enumerate(missing_rows_idx):
            top_k_vals = neighbors[top_k_indices[i], j]
            top_k_weights = grg[i, top_k_indices[i]]
            weight_sum = np.nansum(top_k_weights)
            print(f"Row {row_idx}, Column {j}: Sum of top-k weights = {weight_sum}", flush=True)
            if weight_sum > 0:
                weighted_values = Counter()
                for val, weight in zip(top_k_vals, top_k_weights):
                    if pd.notna(val):
                        weighted_values[val] += weight
                if weighted_values:
                    data[row_idx, j] = weighted_values.most_common(1)[0][0]
            else:
                print(f"Row {row_idx}, Column {j}: Weight sum <= 0 ({weight_sum}), using mode.", flush=True)
                mode_value = df_filled[j].mode()[0] if not df_filled[j].mode().empty else np.nan
                data[row_idx, j] = mode_value
    
    df_filled.iloc[:, :] = data
    return df_filled

def gray_relational_grade_cat_vectorized(matrix, queries, xi=0.5):
    delta = (matrix != queries[:, np.newaxis, :]).astype(float)
    matrix_isna = np.array([[pd.isna(val) for val in row] for row in matrix])
    queries_isna = np.array([[pd.isna(val) for val in row] for row in queries])
    queries_isna_expanded = queries_isna[:, np.newaxis, :]
    nan_mask = matrix_isna | queries_isna_expanded
    delta[nan_mask] = np.nan
    min_delta, max_delta = np.nanmin(delta), np.nanmax(delta)
    
    if np.isnan(max_delta) or max_delta == min_delta:
        grc = np.ones_like(delta) if min_delta == 0 else np.zeros_like(delta)
    else:
        grc = (min_delta + xi * max_delta) / (delta + xi * max_delta)
        grc = np.nan_to_num(grc, nan=0.0)
    
    return np.nanmean(grc, axis=2)

def gray_knn_impute_mixed_vectorized(df, k=5, xi=0.5, batch_size=200):
    df_filled = df.copy()
    df_filled.replace(r'^\s*$', np.nan, regex=True, inplace=True)
    data = df_filled.to_numpy(dtype=object)
    col_types = infer_column_types(df_filled)
    print(f"Column types inferred: {col_types}", flush=True)
    
    for j in range(data.shape[1]):
        missing_mask = pd.isna(data[:, j])
        if not np.any(missing_mask):
            print(f"Column {j}: No missing values, skipping.", flush=True)
            continue
        missing_rows_idx = np.where(missing_mask)[0]
        missing_rows = data[missing_rows_idx, :]
        neighbor_mask = ~pd.isna(data[:, j])
        neighbors = data[neighbor_mask, :]
        if len(neighbors) < k:
            print(f"Column {j}: Insufficient neighbors ({len(neighbors)} < {k}), using mean/mode.", flush=True)
            if col_types[j] == 'numerical':
                column_mean = np.nanmean(df_filled[j])
                data[missing_rows_idx, j] = column_mean if not np.isnan(column_mean) else 0.0
            else:
                mode_value = df_filled[j].mode()[0] if not df_filled[j].mode().empty else np.nan
                data[missing_rows_idx, j] = mode_value
            continue
        n_missing = missing_rows.shape[0]
        batch_size = min(200, max(10, int(n_missing / 10)))
        print(f"Column {j}: {n_missing} missing rows, using batch_size: {batch_size}", flush=True)
        
        grg_all = np.zeros((n_missing, neighbors.shape[0]), dtype=np.float64)
        
        for start in range(0, n_missing, batch_size):
            end = min(start + batch_size, n_missing)
            batch_missing_rows = missing_rows[start:end, :]
            
            delta = np.zeros((batch_missing_rows.shape[0], neighbors.shape[0], data.shape[1]))
            for col, col_type in enumerate(col_types):
                if col_type == 'numerical':
                    m_col = pd.to_numeric(neighbors[:, col], errors='coerce')
                    q_col = pd.to_numeric(batch_missing_rows[:, col], errors='coerce')
                    delta[:, :, col] = np.abs(m_col - q_col[:, np.newaxis])
                else:
                    delta[:, :, col] = (neighbors[:, col] != batch_missing_rows[:, col, np.newaxis]).astype(float)
                    m_isna = np.array([pd.isna(v) for v in neighbors[:, col]])
                    q_isna = np.array([pd.isna(v) for v in batch_missing_rows[:, col]])
                    delta[:, :, col][m_isna | q_isna[:, np.newaxis]] = np.nan
            
            min_delta, max_delta = np.nanmin(delta), np.nanmax(delta)
            if np.isnan(max_delta) or max_delta == min_delta:
                grc = np.ones_like(delta) if min_delta == 0 else np.zeros_like(delta)
            else:
                grc = (min_delta + xi * max_delta) / (delta + xi * max_delta)
                grc = np.nan_to_num(grc, nan=0.0)
            grg_all[start:end, :] = np.nanmean(grc, axis=2)
        
        top_k_indices = np.argsort(grg_all, axis=1)[:, -k:]
        
        for i, row_idx in enumerate(missing_rows_idx):
            top_k_vals = neighbors[top_k_indices[i], j]
            top_k_weights = grg_all[i, top_k_indices[i]]
            weight_sum = np.nansum(top_k_weights)
            print(f"Row {row_idx}, Column {j}: Sum of top-k weights = {weight_sum}", flush=True)
            if weight_sum > 0:
                if col_types[j] == 'numerical':
                    valid_vals = [float(v) for v in top_k_vals if pd.notna(v)]
                    valid_weights = [w for v, w in zip(top_k_vals, top_k_weights) if pd.notna(v)]
                    if valid_weights and sum(valid_weights) > 0:
                        data[row_idx, j] = np.sum([v * w for v, w in zip(valid_vals, valid_weights)]) / sum(valid_weights)
                else:
                    weighted_values = Counter()
                    for val, weight in zip(top_k_vals, top_k_weights):
                        if pd.notna(val):
                            weighted_values[val] += weight
                    if weighted_values:
                        data[row_idx, j] = weighted_values.most_common(1)[0][0]
            else:
                print(f"Row {row_idx}, Column {j}: Weight sum <= 0 ({weight_sum}), using mean/mode.", flush=True)
                if col_types[j] == 'numerical':
                    column_mean = np.nanmean(df_filled[j])
                    data[row_idx, j] = column_mean if not np.isnan(column_mean) else 0.0
                else:
                    mode_value = df_filled[j].mode()[0] if not df_filled[j].mode().empty else np.nan
                    data[row_idx, j] = mode_value
    
    df_filled.iloc[:, :] = data
    return df_filled

def infer_column_types(df):
    col_types = []
    for col in df.columns:
        values = df[col].dropna()
        if values.empty:
            col_types.append('categorical')
        else:
            coerced = pd.to_numeric(values, errors='coerce')
            col_types.append('numerical' if coerced.isna().mean() < 0.5 else 'categorical')
    return col_types

def process_datasets(incomplete_root, original_root, imputed_root, output_excel="Table-NRMS-AE.xlsx"):
    numerical_subfolders = [
        ('Iris', 149), ('Wine', 177), ('Sonar', 207), ('Glass', 213), ('Sheart', 269),
        ('Bupa', 344), ('Ionosphere', 350), ('DERM', 357), ('Difdoug', 399), ('BCW', 682),
        ('PID', 767), ('4-gauss', 799), ('Yeast', 1483), ('CNP', 3999), ('Spam', 4596)
    ]
    categorical_subfolders = [
        ('HOV', 231), ('TTTTEG', 957), ('Splice', 3189), ('MUSH', 5643)
    ]
    mixed_subfolders = [
        ('Aheart', 461), ('CREDIT', 652), ('Abalone', 4176)
    ]
    
    all_subfolders = [
        (subfolder, 'numerical', rows) for subfolder, rows in numerical_subfolders
    ] + [
        (subfolder, 'categorical', rows) for subfolder, rows in categorical_subfolders
    ] + [
        (subfolder, 'mixed', rows) for subfolder, rows in mixed_subfolders
    ]
    
    dataset_types = {
        'numerical': (gray_knn_impute_gpu, {'xi': 0.5}),
        'categorical': (gray_knn_impute_cat_vectorized, {'xi': 0.9}),
        'mixed': (gray_knn_impute_mixed_vectorized, {'xi': 0.5})
    }
    
    if not os.path.exists(imputed_root):
        os.makedirs(imputed_root)
    
    try:
        wb = load_workbook(output_excel)
        ws = wb.active
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file {output_excel} not found.")
    except Exception as e:
        raise Exception(f"Failed to load Excel file: {str(e)}")
    
    dataset_names = [subfolder for subfolder, _, _ in all_subfolders]
    print(f"Processing datasets: {', '.join(dataset_names)}", flush=True)
    
    for subfolder, data_type, rows in all_subfolders:
        if data_type == 'categorical':
            k = max(50, int(math.sqrt(rows) / 2))
        else:
            k = max(5, int(math.sqrt(rows) / 2))
        
        print(f"Subfolder: {subfolder}, Rows: {rows}, Using k: {k}", flush=True)
        
        subfolder_path = os.path.join(incomplete_root, subfolder)
        if not os.path.isdir(subfolder_path):
            print(f"Subfolder not found: {subfolder_path}", flush=True)
            continue
        
        full_file = os.path.join(original_root, f"{subfolder}.xlsx")
        imputed_subfolder = os.path.join(imputed_root, subfolder)
        
        if not os.path.exists(imputed_subfolder):
            os.makedirs(imputed_subfolder)
        
        print(f"Checking file: {full_file}", flush=True)
        if not os.path.exists(full_file):
            print(f"Full file not found: {full_file}", flush=True)
            continue
        dtype = None if data_type == 'numerical' else str
        df_true = pd.read_excel(full_file, header=None, dtype=dtype)
        print(f"Successfully loaded df_true for {subfolder}", flush=True)
        
        impute_func, params = dataset_types[data_type]
        params['k'] = k
        
        for file in os.listdir(subfolder_path):
            if not file.endswith('.xlsx'):
                continue
            
            print(f"Processing file: {file}", flush=True)
            missing_file = os.path.join(subfolder_path, file)
            df_missing = pd.read_excel(missing_file, header=None, dtype=dtype)
            missing_mask = df_missing.isna()
            
            if 'batch_size' in params:
                df_imputed = impute_func(df_missing, **params)
            else:
                df_imputed = impute_func(df_missing, **params)
            
            nrms = 'N/A'
            ae = 'N/A'
            total_missing_cat = 0
            key_params = f"k={k}, xi={params.get('xi', 'N/A')}"
            
            if data_type == 'numerical':
                mask = missing_mask.to_numpy()
                diff = df_imputed.to_numpy()[mask] - df_true.to_numpy()[mask]
                nrms = np.sqrt(np.sum(diff ** 2)) / np.sqrt(np.sum(df_true.to_numpy()[mask] ** 2)) if np.sum(df_true.to_numpy()[mask] ** 2) > 0 else 0
                if nrms != 'N/A' and nrms > 1:
                    nrms = 1.0
                print(f"NRMS for {file}: {nrms}", flush=True)
            elif data_type == 'categorical':
                mask = missing_mask.to_numpy()
                imputed_values = df_imputed.to_numpy()[mask]
                true_values = df_true.to_numpy()[mask]
                correct_matches = np.sum((imputed_values == true_values) & mask[mask])
                total_missing_cat = np.sum(mask)
                ae = correct_matches / total_missing_cat if total_missing_cat > 0 else 0
                print(f"AE for {file}: {ae}", flush=True)
            else:
                col_types = infer_column_types(df_missing)
                mask = missing_mask.to_numpy()
                numerical_cols = [i for i, t in enumerate(col_types) if t == 'numerical']
                categorical_cols = [i for i, t in enumerate(col_types) if t == 'categorical']
                
                if numerical_cols:
                    num_mask = mask[:, numerical_cols]
                    imputed_numeric = df_imputed.iloc[:, numerical_cols].apply(lambda x: pd.to_numeric(x, errors='coerce')).to_numpy()
                    true_numeric = df_true.iloc[:, numerical_cols].apply(lambda x: pd.to_numeric(x, errors='coerce')).to_numpy()
                    diff = imputed_numeric[num_mask] - true_numeric[num_mask]
                    nrms_num = np.sum(diff ** 2)
                    nrms_denom = np.sum(true_numeric[num_mask] ** 2)
                    nrms = np.sqrt(nrms_num) / np.sqrt(nrms_denom) if nrms_denom > 0 else 0
                    if nrms != 'N/A' and nrms > 1:
                        nrms = 1.0
                    print(f"NRMS for {file}: {nrms}", flush=True)
                if categorical_cols:
                    cat_mask = mask[:, categorical_cols]
                    imputed_values = df_imputed.iloc[:, categorical_cols].to_numpy()[cat_mask]
                    true_values = df_true.iloc[:, categorical_cols].to_numpy()[cat_mask]
                    correct_matches = np.sum((imputed_values == true_values) & cat_mask[cat_mask])
                    total_missing_cat = np.sum(cat_mask)
                    ae = correct_matches / total_missing_cat if total_missing_cat > 0 else 0
                    print(f"AE for {file}: {ae}", flush=True)
            
            dataset_name = file.split('.')[0]
            
            updated = False
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
                if row[0].value == dataset_name:
                    if row[0].value is None:
                        print(f"Debug: Cell A{row[0].row} is empty or None, skipping update.", flush=True)
                        continue
                    try:
                        row[1].value = nrms if nrms != 'N/A' else ''
                        row[1].number_format = 'General'
                        row[2].value = ae if ae != 'N/A' else ''
                        row[2].number_format = 'General'
                        row[3].value = key_params
                        updated = True
                        print(f"Updated Excel row for {dataset_name}", flush=True)
                        wb.save(output_excel)
                    except PermissionError:
                        print(f"Error: Permission denied when updating row {row[0].row} for {dataset_name}. Close the excel file", flush=True)
                        continue
                    except Exception as e:
                        print(f"Error updating row {row[0].row} for {dataset_name}: {str(e)}", flush=True)
                        continue
                    break
            else:
                print(f"Warning: Dataset '{dataset_name}' not found in Excel file. Checking possible mismatches:", flush=True)
                datasets_in_file = [cell.value for cell in ws['A'][1:] if cell.value is not None]
                print(f"Datasets in file: {datasets_in_file}", flush=True)
                print(f"Current dataset_name: '{dataset_name}' (type: {type(dataset_name)})", flush=True)
            
            imputed_file = os.path.join(imputed_subfolder, f"imputed_{file}")
            df_imputed.to_excel(imputed_file, index=False, header=False)
            print(f"Saved imputed file: {imputed_file}", flush=True)
    
    try:
        wb.save(output_excel)
        wb_check = load_workbook(output_excel)
        ws_check = wb_check.active
        updated_nrms = any(cell.value is not None and cell.row > 1 for cell in ws_check['B'] if cell.value != '')
        if not updated_nrms:
            raise Exception(f"Final update failed: No changes detected in 'NRMS' column after saving {output_excel}.")
        else:
            print(f"Excel file successfully updated: {output_excel}", flush=True)
    except PermissionError:
        raise Exception(f"Error: Permission denied when saving {output_excel}.")
    except Exception as e:
        raise Exception(f"Error saving Excel file: {str(e)}. Update may not have been applied.")

incomplete_root = r"Incomplete Datasets Without Labels"
original_root = r"Original Datasets Without Labels"
imputed_root = r"Imputed Datasets"

process_datasets(incomplete_root, original_root, imputed_root, output_excel="Table-NRMS-AE.xlsx")